from openpyxl import *
from openpyxl.styles import *

mainWorkbook = load_workbook(filename= '11-23-2020-UDT-LENOVO-DATA FILE Batch 1.xlsx') # loads the mainWorkbook that needs to be copied
outputWorkbook = load_workbook(filename='UDT PARTNER TEMPLATE 12-24-20.xlsx') # loads the outputWorkbook that needs

def retrieveClaimNumbers(batchSheet): # opens and copies claim numbers from mainworkbook
    claimNumbers = [] # claimNumbers empty array which will be store a value
    for i in range(2,batchSheet.max_row+1): # this loops start at row 2 and iterates down to the max value
        cellValue = batchSheet.cell(row = i,column =8).value # udt claim numbers = 8
        claimNumbers.append(cellValue) # takes the iterated row cell value and appends the array
    return claimNumbers # returns the value of back into the function

def pasteClaimNumbers(claimNumbers, sheetName): # takes in the input from claimnumbers from def retrieveClaimNumbers which is appended
    sheet = outputWorkbook[sheetName]
    for c in range(0,len(claimNumbers)):
        sheet.cell(row = c+2, column = 9).value=claimNumbers[c] # claim numbers = 9
        for row in outputWorkbook[sheetName]:
            for cell in row:
                cell.alignment = Alignment(horizontal = None, wrap_text=True, vertical='bottom')
            
print('Work, work')
print("Copying over claim numbers info to output workbook")
for i, sheet in enumerate(mainWorkbook.sheetnames):
    currentSheet = mainWorkbook[sheet]
    outputSheet = outputWorkbook.sheetnames[i]
    claimNumbers = retrieveClaimNumbers(currentSheet)
    pasteClaimNumbers(claimNumbers, outputSheet)

#==============================================================================================================================================
def retrieveTrackingNumbers(batchSheet): # retrieves and copies the tracking number from main
    outboundNumbers = [] # stores outbound numbers from mainWorkbook
    for i in range(2,batchSheet.max_row+1):
        cellValue = batchSheet.cell(row = i,column =9).value # outbound = 9
        outboundNumbers.append(cellValue)
    return outboundNumbers

def pasteTrackingNumbers(outboundNumbers, sheetName):
    sheet = outputWorkbook[sheetName]
    for column in sheet['G']:
        column.number_format = '0000'
    for c in range(0,len(outboundNumbers)):
        sheet.cell(row = c+2, column = 7).value=outboundNumbers[c] # tracking number = 7

print("Copying over tracking number info to output workbook")
for i, sheet in enumerate(mainWorkbook.sheetnames):
    currentSheet = mainWorkbook[sheet]
    outputSheet = outputWorkbook.sheetnames[i]
    outboundNumbers = retrieveTrackingNumbers(currentSheet)
    pasteTrackingNumbers(outboundNumbers, outputSheet)
#==============================================================================================================================================
def retrieveResolution(batchSheet):
    resolutionStatus = []
    for i in range(2,batchSheet.max_row+1):
        cellValue = batchSheet.cell(row = i,column =7).value # resolution = 7
        resolutionStatus.append(cellValue)
    return resolutionStatus

def pasteResolution(resolutionStatus, sheetName):
    sheet = outputWorkbook[sheetName]
    for c in range(0,len(resolutionStatus)):
        sheet.cell(row = c+2, column = 4).value=resolutionStatus[c] # Status = 4

print("Copying over resolution info to output workbook")
for i , sheet in enumerate(mainWorkbook.sheetnames):
    currentSheet = mainWorkbook[sheet]
    outputSheet = outputWorkbook.sheetnames[i]
    resolutionStatus = retrieveResolution(currentSheet)
    pasteResolution(resolutionStatus, outputSheet)
#==============================================================================================================================================
def retrieveShippingDate(batchSheet):
    shippedDate = []
    for i in range(2,batchSheet.max_row+1):
        cellValue = batchSheet.cell(row = i,column =10).value # Date returned = 10
        shippedDate.append(cellValue)
    return shippedDate

def pasteShippingDate(shippedDate, sheetName):
    sheet = outputWorkbook[sheetName]
    for column in sheet['F']:
        column.number_format = 'mm/dd/yyyy'    
    for c in range(0,len(shippedDate)):
        sheet.cell(row = c+2, column = 6).value=shippedDate[c] # Shipped Date = 6

print("Copying over shipping date info to output workbook")
for i , sheet in enumerate(mainWorkbook.sheetnames):
    currentSheet = mainWorkbook[sheet]
    outputSheet = outputWorkbook.sheetnames[i]
    shippedDate = retrieveShippingDate(currentSheet)
    pasteShippingDate(shippedDate, outputSheet)
#==============================================================================================================================================
outputWorkbook.save('UDT PARTNER TEMPLATE 12-24-20.xlsx')
print('Jobs done!')
